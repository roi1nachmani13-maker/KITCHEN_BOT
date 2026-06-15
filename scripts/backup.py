#!/usr/bin/env python3
"""
Manual backup script: exports all sheets to a local Excel file.
Can also be triggered from cron or docker.

Run:
    python scripts/backup.py [output_dir]
"""
from __future__ import annotations

import sys
import os
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from src.sheets.client import sheets_client, SHEET_INVENTORY, SHEET_COMPLETIONS, SHEET_HISTORY
from src.utils.config import config
from src.utils.logger import get_logger

log = get_logger("backup")


def export_to_excel(output_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")

    for sheet_name in [SHEET_INVENTORY, SHEET_COMPLETIONS, SHEET_HISTORY]:
        try:
            records = sheets_client.get_all_records(sheet_name)
            ws = wb.create_sheet(title=sheet_name)

            if not records:
                ws.append([f"No data in {sheet_name}"])
                continue

            # Headers
            headers = list(records[0].keys())
            ws.append(headers)
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for record in records:
                ws.append([record.get(h, "") for h in headers])

            # Auto column width
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            log.info(f"Exported {len(records)} rows from '{sheet_name}'")
        except Exception as e:
            log.error(f"Error exporting {sheet_name}: {e}")

    wb.save(output_path)
    log.info(f"Backup saved: {output_path}")


def main() -> None:
    output_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("backups")
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"kitchen_backup_{timestamp}.xlsx"

    config.validate()
    sheets_client.connect()

    print(f"📦 Exporting to: {output_path}")
    export_to_excel(output_path)
    print(f"✅ Backup complete: {output_path}")


if __name__ == "__main__":
    main()
